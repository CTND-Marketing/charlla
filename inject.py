import csv, io, json, re
from openpyxl import load_workbook
from datetime import datetime

def read_euckr(path):
    with open(path, 'rb') as f:
        return f.read().decode('euc-kr', errors='replace')

# ── config.json 읽기
try:
    with open('data/config.json', encoding='utf-8') as f:
        config = json.load(f)
except:
    config = {}

last_updated    = config.get('lastUpdated', '2026-04-29')
weekly_visitors = config.get('weeklyVisitors', [None, None, None, None])
ad_costs        = config.get('adCosts', {'google':250,'gdn':100,'naver':210,'cafe24':369})
ga4_cumulative  = config.get('ga4Cumulative', [0, 0, 0, 0])
months_data     = config.get('months', {})  # 월별 저장 데이터

try:
    report_date  = datetime.fromisoformat(last_updated)
    report_month = report_date.month
    report_year  = report_date.year
    report_day   = report_date.day
except:
    report_date  = datetime(2026, 4, 29)
    report_month = 4
    report_year  = 2026
    report_day   = 29

cur_month_key = f'{report_year}-{report_month:02d}'  # e.g. "2026-04"

# ── 채널/이벤트 매핑
def map_visitor(g, s, m):
    g,s,m = g.strip().lower(), s.strip().lower(), m.strip().lower()
    if g == 'direct' or (s == '(direct)' and m == '(none)'): return ('direct','direct')
    if m in ('brand_search_pc','brand_search_m') or s == 'bs': return ('paid search','brand_Search')
    if g in ('paid search','paid video') and s == 'google' and m == 'cpc': return ('paid search','google')
    if g in ('paid search','paid shopping') and 'naver' in s: return ('paid search','naver')
    if s in ('ig','instagram','l.instagram.com'): return ('paid search','instagram')
    if g == 'organic search' and s == 'google': return ('organic search','google')
    if g == 'organic search' and ('naver' in s or s in ('m.naver.com','m.search.naver.com')): return ('organic search','naver')
    if g == 'organic search' and s == 'bing': return ('organic search','etc.(bing)')
    if s in ('inblog','blog','charlla_inblog','blog.naver.com','m.blog.naver.com','blog.catenoid.net'): return ('unassigned','blog')
    if g == 'organic social': return ('unassigned','blog')
    if s == 'iboss': return ('unassigned','iboss')
    if s == 'stibee' or m in ('ebook','ebook2'): return ('unassigned','stibee')
    if s == 'openads': return ('unassigned','openads')
    if s in ('catenoid.net','charlla.io'): return ('referral','catenoid.net / charlla.io')
    if s in ('chatgpt.com','gemini.google.com') or (g == 'referral' and s == 'ai'): return ('referral','ai')
    if g == 'referral' and any(x in s for x in ('cafe24','makeshop','sixshop')): return ('referral','cafe24 (ad)')
    if s in ('gdn','viimstudio','(data not available)') or (g == 'display' and s == 'google') or g == 'cross-network': return ('display','GDN banner')
    if 'cafe24' in s and m == 'floating_banner': return ('display','CAFE24 banner')
    if g == 'display' and ('cafe24' in s or m in ('banner','floating_banner')): return ('display','CAFE24 banner')
    if g == 'display': return ('display','etc.(GFA, DMP..)')
    if g == 'referral': return ('referral','etc.')
    if g == 'unassigned': return ('unassigned','etc.')
    return None

def map_event(s, m):
    s,m = s.strip().lower(), m.strip().lower()
    if s == '(direct)' or m == '(none)': return ('direct','direct')
    if m in ('brand_search_pc','brand_search_m') or s == 'bs': return ('paid search','brand_Search')
    if s == 'google' and m == 'cpc': return ('paid search','google')
    if s == 'naver' and m == 'cpc': return ('paid search','naver')
    if s in ('ig','instagram'): return ('paid search','instagram')
    if s == 'google' and m == 'organic': return ('organic search','google')
    if ('naver' in s or s == 'm.search.naver.com') and m in ('organic','referral'): return ('organic search','naver')
    if s == 'bing' and m == 'organic': return ('organic search','etc.(bing)')
    if s in ('blog.naver.com','m.blog.naver.com','inblog'): return ('unassigned','blog')
    if s in ('gdn','viimstudio','(data not available)'): return ('display','GDN banner')
    if 'cafe24' in s and m in ('referral','banner','floating_banner'): return ('referral','cafe24 (ad)')
    if s in ('chatgpt.com','gemini.google.com'): return ('referral','ai')
    if s in ('accounts.google.com','mail.google.com'): return ('direct','direct')
    if m == 'referral': return ('referral','etc.')
    if m in ('ebook','ebook2'): return ('unassigned','stibee')
    return None

stage_map = {
    'ga4-btn-home-freetrial':'ft','ga4-btn-gnb-freetrial':'ft','ga4-btn-basic-free':'ft',
    'ga4_btn_basic_free':'ft','ga4-btn-home-footer-freetrial':'ft','ga4-btn-lite-free':'ft',
    'ga4-btn-multiplayer-free':'ft','ga4-btn-playersetting-footer-free':'ft',
    'ga4-btn-playersetting-free':'ft','ga4-btn-playlist-footer-free':'ft',
    'ga4-btn-playlist-free':'ft','ga4-btn-playlist-free1':'ft','ga4-btn-playlist-free2':'ft',
    'ga4-btn-standard-free':'ft','ga4-btn-statistic-free':'ft','ga4-btn-statistic-footer-free':'ft',
    'sign_up_create_account':'ac','sign-up-create-account-social':'ac',
    'ga4-btn-policy-next':'po','ga4-btn-policy-social-next':'po',
    'sign_up':'su',
    'ga4-btn-shopinfo-next':None,'start_sign_up':None,
}

structure = [
    ('direct',         ['direct']),
    ('unassigned',     ['blog','outstanding','qletter','iboss','openads','stibee','etc.']),
    ('paid search',    ['brand_Search','google','naver','instagram']),
    ('organic search', ['google','naver','etc.(bing)']),
    ('referral',       ['catenoid.net / charlla.io','cafe24 (ad)','makeshop','ai','etc.']),
    ('display',        ['GDN banner','CAFE24 banner','IBOSS banner','Outstanding banner','etc.(GFA, DMP..)']),
    ('SNS',            ['facebook / linkedin']),
]

totals = {cat+'|'+row: {'v':0,'ft':0,'ac':0,'po':0,'su':0} for cat,rows in structure for row in rows}

# ── CSV 파싱
for row in list(csv.reader(io.StringIO(read_euckr('data/visitors.csv'))))[1:]:
    if len(row) < 4: continue
    try: n = int(row[3].strip())
    except: continue
    r = map_visitor(row[0],row[1],row[2])
    if r and r[0]+'|'+r[1] in totals: totals[r[0]+'|'+r[1]]['v'] += n

for row in list(csv.reader(io.StringIO(read_euckr('data/events.csv'))))[1:]:
    if len(row) < 4: continue
    try: n = int(row[3].strip())
    except: continue
    st = stage_map.get(row[0].strip())
    if not st: continue
    r = map_event(row[1],row[2])
    if r and r[0]+'|'+r[1] in totals: totals[r[0]+'|'+r[1]][st] += n

# ── Metabase xlsx
all_month_su   = {}
all_month_paid = {}
week_mb_data   = {1:0, 2:0, 3:0, 4:0}

try:
    wb = load_workbook('data/metabase.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows_mb = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows_mb[0]]
    join_idx = headers.index('가입일') if '가입일' in headers else None
    paid_idx = headers.index('유료전환 여부') if '유료전환 여부' in headers else None
    for row in rows_mb[1:]:
        if join_idx is None: continue
        raw = row[join_idx]
        if not raw: continue
        d = raw if isinstance(raw, datetime) else None
        if not d:
            try: d = datetime.fromisoformat(str(raw))
            except: continue
        mo = d.month
        yr = d.year
        key = f'{yr}-{mo:02d}'
        all_month_su[key]   = all_month_su.get(key, 0) + 1
        if paid_idx is not None and str(row[paid_idx]).strip().upper() == 'Y':
            all_month_paid[key] = all_month_paid.get(key, 0) + 1
        # 당월 주차 분기
        if mo == report_month and yr == report_year:
            if d.day <= 8:    week_mb_data[1] += 1
            elif d.day <= 15: week_mb_data[2] += 1
            elif d.day <= 22: week_mb_data[3] += 1
            else:             week_mb_data[4] += 1
except Exception as e:
    print(f"Metabase 처리 오류: {e}")

# ── 집계
all_data = [(cat, row_name, totals[cat+'|'+row_name]) for cat, rows in structure for row_name in rows]
total_v  = sum(d['v']  for _,_,d in all_data)
total_ft = sum(d['ft'] for _,_,d in all_data)
total_ac = sum(d['ac'] for _,_,d in all_data)
total_po = sum(d['po'] for _,_,d in all_data)
total_su = sum(d['su'] for _,_,d in all_data)
ga4_cvr  = round(total_su/total_v*100, 2) if total_v else 0
mb_cur   = all_month_su.get(cur_month_key, 0)
mb_cvr   = round(mb_cur/total_v*100, 2) if total_v else 0

# ── GA4 주차별 자동 계산
if report_day <= 8:    cur_week = 1
elif report_day <= 15: cur_week = 2
elif report_day <= 22: cur_week = 3
else:                  cur_week = 4

prev_cum = ga4_cumulative[cur_week - 2] if cur_week >= 2 else 0
week_ga4 = [None, None, None, None]
for w in range(1, cur_week):
    p = ga4_cumulative[w - 2] if w >= 2 else 0
    c_val = ga4_cumulative[w - 1]
    week_ga4[w - 1] = c_val - p if c_val > 0 else None
week_ga4[cur_week - 1] = total_su - prev_cum

# config.json에 GA4 누적값 저장
new_cum = ga4_cumulative[:]
new_cum[cur_week - 1] = total_su
config['ga4Cumulative'] = new_cum

# ── 채널 집계
ch_defs = [
    ('검색/배너광고','#3b82f6', lambda ct,n: ct=='paid search' or n=='GDN banner'),
    ('직접 유입',    '#10b981', lambda ct,n: ct=='direct'),
    ('카페24',       '#f59e0b', lambda ct,n: n in ('cafe24 (ad)','CAFE24 banner')),
    ('자연유입',     '#06b6d4', lambda ct,n: ct=='organic search' or n in ('blog','ai')),
    ('기타',         '#cbd5e1', lambda ct,n: (ct in ('unassigned','SNS')) or (ct=='referral' and n not in ('cafe24 (ad)',)) or (ct=='display' and n!='GDN banner')),
]
ch_data = []
for name, color, fn in ch_defs:
    rows = [(ct,n,d) for ct,n,d in all_data if fn(ct,n)]
    v  = sum(d['v']  for _,_,d in rows)
    su = sum(d['su'] for _,_,d in rows)
    ch_data.append({'name':name,'v':v,'su':su,'color':color})

# ── adEff
ae_defs = [
    ("Google\nKeyword", 'paid search', ['google'],                     ad_costs.get('google',250)*10000, '#3b82f6'),
    ("GDN",             'display',     ['GDN banner'],                 ad_costs.get('gdn',100)*10000,    '#7c3aed'),
    ("Naver\n검색광고", 'paid search', ['naver','brand_Search'],       ad_costs.get('naver',210)*10000,  '#10b981'),
    ("카페24\n배너",    None,          ['cafe24 (ad)','CAFE24 banner'], ad_costs.get('cafe24',369)*10000, '#f59e0b'),
]
adEff = []
for name, cat, rnames, cost, color in ae_defs:
    rows = [(ct,n,d) for ct,n,d in all_data if (cat is None or ct==cat) and n in rnames]
    v  = sum(d['v']  for _,_,d in rows)
    su = sum(d['su'] for _,_,d in rows)
    adEff.append({'name':name,'v':v,'su':su,'cost':cost,'color':color})

# ── 현재 월 데이터를 months에 저장
def cvr_type(v, su):
    if v == 0: return 'null'
    r = su/v*100
    if r >= 6: return 'high'
    if r >= 3: return 'mid'
    if r > 0: return 'low'
    return 'null'

td = {a+'|'+b: d for a,b,d in all_data}

raw_rows = {}
for cat, rows in structure:
    raw_rows[cat] = {}
    for row_name in rows:
        d = td.get(cat+'|'+row_name, {'v':0,'ft':0,'ac':0,'po':0,'su':0})
        raw_rows[cat][row_name] = dict(d)

cvrS = sorted(ch_data, key=lambda d: -(d['su']/d['v']) if d['v'] else 0)
for d in cvrS:
    d['cvr'] = round(d['su']/d['v']*100, 1) if d['v'] else 0

wv = weekly_visitors + [None] * (4 - len(weekly_visitors))
w1 = wv[0] if wv[0] is not None else 0
w2 = wv[1] if wv[1] is not None else 0
w3 = wv[2] if wv[2] is not None else 0
w4 = wv[3]
week_v = [w1, w2, w3, w4]
week_mb = [week_mb_data[1], week_mb_data[2], week_mb_data[3],
           week_mb_data[4] if week_mb_data[4] > 0 else None]

# 채널 비중 추이 (이전 월 데이터에서 가져오기)
def get_prev_months(cur_key, n=3):
    yr, mo = int(cur_key[:4]), int(cur_key[5:])
    result = []
    for _ in range(n):
        mo -= 1
        if mo == 0: mo = 12; yr -= 1
        result.insert(0, f'{yr}-{mo:02d}')
    return result

prev_keys = get_prev_months(cur_month_key, 3)
month_labels_list = []
for k in prev_keys:
    mo = int(k[5:])
    month_labels_list.append(f'{mo}월')
month_labels_list.append(f'{report_month}월')

# 채널 비중 추이 데이터 (이전 월은 저장된 데이터에서)
ch_trend_pct = [[] for _ in range(5)]
ch_trend_abs = [[] for _ in range(5)]
for k in prev_keys:
    saved = months_data.get(k, {})
    saved_ch = saved.get('channels', [])
    saved_su = saved.get('totalSu', 0)
    for i in range(5):
        if i < len(saved_ch) and saved_su > 0:
            ch_trend_abs[i].append(saved_ch[i].get('su', 0))
            ch_trend_pct[i].append(round(saved_ch[i].get('su', 0)/saved_su*100, 1))
        else:
            ch_trend_abs[i].append(0)
            ch_trend_pct[i].append(0)
# 현재 월 추가
cur_total_su_ch = sum(d['su'] for d in ch_data)
for i, d in enumerate(ch_data):
    ch_trend_abs[i].append(d['su'])
    ch_trend_pct[i].append(round(d['su']/cur_total_su_ch*100, 1) if cur_total_su_ch else 0)

# 월별 차트 데이터
monthly_v_list   = []
monthly_ga4_list = []
monthly_mb_list  = []
monthly_ga4_cvr_list = []
monthly_mb_cvr_list  = []
monthly_paid_list    = []
monthly_paid_mb_list = []
monthly_paid_r_list  = []
for k in prev_keys:
    saved = months_data.get(k, {})
    sv = saved.get('totalV', 0)
    monthly_v_list.append(sv)
    monthly_ga4_list.append(saved.get('totalSu', 0))
    monthly_mb_list.append(saved.get('mbSu', 0))
    monthly_ga4_cvr_list.append(saved.get('ga4Cvr', 0))
    monthly_mb_cvr_list.append(saved.get('mbCvr', 0))
    monthly_paid_list.append(saved.get('paidSu', 0))
    monthly_paid_mb_list.append(saved.get('mbSu', 0))
    monthly_paid_r_list.append(saved.get('paidRate', 0))

monthly_v_list.append(total_v)
monthly_ga4_list.append(total_su)
monthly_mb_list.append(mb_cur)
monthly_ga4_cvr_list.append(ga4_cvr)
monthly_mb_cvr_list.append(mb_cvr)
monthly_paid_list.append(all_month_paid.get(cur_month_key, 0))
monthly_paid_mb_list.append(mb_cur)
paid_cur = all_month_paid.get(cur_month_key, 0)
monthly_paid_r_list.append(round(paid_cur/mb_cur*100, 1) if mb_cur else 0)

# ── 현재 월 데이터 months에 저장
months_data[cur_month_key] = {
    'lastUpdated':  last_updated,
    'totalV':       total_v,
    'totalSu':      total_su,
    'ga4Cvr':       ga4_cvr,
    'mbSu':         mb_cur,
    'mbCvr':        mb_cvr,
    'paidSu':       paid_cur,
    'paidRate':     round(paid_cur/mb_cur*100, 1) if mb_cur else 0,
    'weekV':        week_v,
    'weekGa4':      week_ga4,
    'weekMb':       week_mb,
    'channels':     ch_data,
    'adEff':        [{'name':d['name'],'v':d['v'],'su':d['su'],'cost':d['cost'],'color':d['color']} for d in adEff],
    'rawRows':      raw_rows,
    'totalFt':      total_ft,
    'totalAc':      total_ac,
    'totalPo':      total_po,
    'ga4Cumulative': new_cum,
}
config['months'] = months_data
config['ga4Cumulative'] = new_cum

try:
    with open('data/config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"config.json 저장 완료 ({cur_month_key})")
except Exception as e:
    print(f"config.json 저장 실패: {e}")

# ── rawData JS 생성
raw_lines = ['[']
for cat, rows in structure:
    raw_lines.append("  { cat: '" + cat + "', rows: [")
    for row_name in rows:
        d = td.get(cat+'|'+row_name, {'v':0,'ft':0,'ac':0,'po':0,'su':0})
        v,ft,ac,po,su = d['v'],d['ft'],d['ac'],d['po'],d['su']
        cvr_val = str(round(su/v*100,1))+'%' if v and su else ('0%' if v else '-')
        raw_lines.append("    { name: '"+row_name+"', v:"+str(v)+", ft:"+str(ft)+", ac:"+str(ac)+", po:"+str(po)+", su:"+str(su)+", cvr:'"+cvr_val+"', cvrType:'"+cvr_type(v,su)+"' },")
    raw_lines.append('  ]},')
raw_lines.append(']')

# ── HTML 읽기
with open('index_template.html', encoding='utf-8') as f:
    c = f.read()

# ── 데이터 주입
raw_start = c.find('const rawData = [')
raw_end   = c.find('const cvrColors', raw_start)
c = c[:raw_start] + 'const rawData = ' + '\n'.join(raw_lines) + ';\n' + c[raw_end:]

funnel = [total_v, total_ft, total_ac, total_po, total_su]
fd_start = c.find('const funnelData = [')
fd_end   = c.find(']', fd_start) + 1
c = c[:fd_start] + 'const funnelData = ' + json.dumps(funnel) + c[fd_end:]

ch_start = c.find('const channels = [')
ch_end   = c.find('const totalSU', ch_start)
ch_lines = ['const channels = [']
for d in ch_data:
    ch_lines.append("  { name:'"+d['name']+"', su:"+str(d['su'])+", v:"+str(d['v'])+", color:'"+d['color']+"' },")
ch_lines.append('];')
c = c[:ch_start] + '\n'.join(ch_lines) + '\n' + c[ch_end:]

tsu_s = c.find('const totalSU = ')
tsu_e = c.find(';', tsu_s) + 1
c = c[:tsu_s] + 'const totalSU = ' + str(total_su) + ';' + c[tsu_e:]

ae_str = 'const adEff = ['
for d in adEff:
    ae_str += "{name:'"+d['name'].replace('\n','\\n')+"',v:"+str(d['v'])+",su:"+str(d['su'])+",cost:"+str(d['cost'])+",color:'"+d['color']+"'},"
ae_str += '];'
ae_s = c.find('const adEff = [')
ae_e = c.find('];', ae_s) + 2
c = c[:ae_s] + ae_str + c[ae_e:]

def rep(html, id_, val):
    return re.sub('id="'+id_+'">[^<]*<', 'id="'+id_+'">'+str(val)+'<', html)

c = rep(c,'kpiTotalV', f'{total_v:,}명')
c = rep(c,'kpiGA4Su',  f'{total_su:,}명')
c = rep(c,'kpiGA4Cvr', str(ga4_cvr)+'%')
c = rep(c,'kpiMBSu',   f'{mb_cur:,}명')
c = rep(c,'kpiMBCvr',  str(mb_cvr)+'%')
c = rep(c,'accTotalV',  str(total_v))
c = rep(c,'accTotalFt', str(total_ft))
c = rep(c,'accTotalAc', str(total_ac))
c = rep(c,'accTotalPo', str(total_po))
c = rep(c,'accTotalSu', str(total_su))
c = rep(c,'accTotalCvr', str(ga4_cvr)+'%')

for i, k in enumerate(['google','gdn','naver','cafe24']):
    v  = adEff[i]['v']
    su = adEff[i]['su']
    cost = adEff[i]['cost']
    cvr_ = round(su/v*100,2) if v else 0
    cpa_ = round(cost/su/10000) if su else 0
    c = rep(c,'kpi_v_'+k,   str(v))
    c = rep(c,'kpi_su_'+k,  str(su))
    c = rep(c,'kpi_cvr_'+k, str(cvr_)+'%')
    c = rep(c,'kpi_cpa_'+k, str(cpa_)+'만')

def drop(a,b): return round((1-b/a)*100,1) if a else 0
c = re.sub(r'id="funnelDrop0">[^<]*<','id="funnelDrop0">'+str(drop(total_v,total_ft))+'<',c)
c = re.sub(r'id="funnelDrop1">[^<]*<','id="funnelDrop1">'+str(drop(total_ft,total_ac))+'<',c)
c = re.sub(r'id="funnelDrop2">[^<]*<','id="funnelDrop2">'+str(drop(total_ac,total_po))+'<',c)
c = re.sub(r'id="funnelDrop3">[^<]*<','id="funnelDrop3">'+str(drop(total_po,total_su))+'<',c)

for wid, wval in [('wv1', w1), ('wv2', w2), ('wv3', w3), ('wv4', '' if w4 is None else w4)]:
    c = re.sub('id="'+wid+'" value="[^"]*"', 'id="'+wid+'" value="'+str(wval)+'"', c)

date_label = last_updated + ' 기준'
c = re.sub('id="lastUpdatedLabel"[^>]*>[^<]*<', 'id="lastUpdatedLabel" class="text-xs text-gray-400">'+date_label+'<', c)

# ── 월별 탭 HTML 생성
def make_tab_html(months_data, cur_key):
    sorted_keys = sorted(months_data.keys())
    if cur_key not in sorted_keys:
        sorted_keys.append(cur_key)
    sorted_keys = sorted(sorted_keys)

    mo_names = {1:'1월',2:'2월',3:'3월',4:'4월',5:'5월',6:'6월',
                7:'7월',8:'8월',9:'9월',10:'10월',11:'11월',12:'12월'}

    tabs_html = '<div id="monthTabBar" class="flex gap-1 mb-6" style="border-bottom:1px solid #e5e7eb;padding-bottom:0;">'
    for k in sorted_keys:
        mo = int(k[5:])
        label = mo_names.get(mo, f'{mo}월')
        is_cur = (k == cur_key)
        active_cls = 'border-b-2 border-blue-500 text-blue-600 font-semibold' if is_cur else 'text-gray-400 hover:text-gray-600'
        badge = '<span class="ml-1 text-xs bg-blue-100 text-blue-600 px-1.5 py-0.5 rounded-full">최신</span>' if is_cur else ''
        tabs_html += f'<button onclick="switchMonth(\'{k}\')" class="tab-btn px-4 py-2 text-sm {active_cls} cursor-pointer" data-month="{k}">{label}{badge}</button>'
    tabs_html += '</div>'

    # 각 월 데이터를 JS 객체로 직렬화
    all_months_js = json.dumps(months_data, ensure_ascii=False)

    switch_script = f'''
<script>
var _allMonths = {all_months_js};
var _curMonth = '{cur_key}';
var _monthNames = {json.dumps({k: mo_names.get(int(k[5:]),k[5:]) for k in sorted_keys}, ensure_ascii=False)};

function switchMonth(key) {{
  if (key === _curMonth) return;
  var d = _allMonths[key];
  if (!d) return;

  // 탭 UI 업데이트
  document.querySelectorAll('.tab-btn').forEach(function(btn) {{
    var isActive = btn.dataset.month === key;
    btn.className = 'tab-btn px-4 py-2 text-sm cursor-pointer ' + (isActive ? 'border-b-2 border-blue-500 text-blue-600 font-semibold' : 'text-gray-400 hover:text-gray-600');
    var badge = btn.querySelector('span');
    if (badge) btn.removeChild(badge);
  }});

  var setEl = function(id, val) {{ var el = document.getElementById(id); if (el) el.textContent = val; }};

  // KPI 업데이트
  setEl('kpiTotalV', (d.totalV||0).toLocaleString() + '명');
  setEl('kpiGA4Su',  (d.totalSu||0).toLocaleString() + '명');
  setEl('kpiGA4Cvr', (d.ga4Cvr||0) + '%');
  setEl('kpiMBSu',   (d.mbSu||0).toLocaleString() + '명');
  setEl('kpiMBCvr',  (d.mbCvr||0) + '%');
  setEl('lastUpdatedLabel', (d.lastUpdated||key) + ' 기준');

  // 주차별 차트
  var mo = parseInt(key.substring(5));
  var wl = [[mo+'월 1주차','('+mo+'/1~'+mo+'/8)'],[mo+'월 2주차','('+mo+'/9~'+mo+'/15)'],[mo+'월 3주차','('+mo+'/16~'+mo+'/22)'],[mo+'월 4주차','('+mo+'/23~'+mo+'/말)']];
  if (window.weeklyAprilChartRef) {{
    var w = window.weeklyAprilChartRef;
    w.data.labels = wl;
    w.data.datasets[0].data = d.weekV || [null,null,null,null];
    w.data.datasets[1].data = d.weekMb || [null,null,null,null];
    w.data.datasets[2].data = d.weekGa4 || [null,null,null,null];
    w.update();
  }}

  // 채널 전환율 차트
  if (window.cvrChartRef && d.channels) {{
    var cvrS2 = d.channels.slice().sort(function(a,b){{ return (b.su/b.v||0)-(a.su/a.v||0); }});
    var cv = window.cvrChartRef;
    cv.data.labels = cvrS2.map(function(x){{return x.name;}});
    cv.data.datasets[0].data = cvrS2.map(function(x){{return x.v?+(x.su/x.v*100).toFixed(1):0;}});
    cv.data.datasets[0].backgroundColor = cvrS2.map(function(x){{return x.color+'bb';}});
    cv.data.datasets[0].borderColor = cvrS2.map(function(x){{return x.color;}});
    cvrSorted.length=0; cvrS2.forEach(function(x){{x.cvr=x.v?+(x.su/x.v*100).toFixed(1):0; cvrSorted.push(x);}});
    cv.update();
  }}

  // 광고 KPI 카드
  if (d.adEff) {{
    var adKeys = ['google','gdn','naver','cafe24'];
    d.adEff.forEach(function(ae, i) {{
      var k2 = adKeys[i];
      var cvr2 = ae.v ? (ae.su/ae.v*100).toFixed(2) : 0;
      var cpa2 = ae.su ? Math.round(ae.cost/ae.su/10000) : 0;
      setEl('kpi_v_'+k2,   ae.v.toLocaleString());
      setEl('kpi_su_'+k2,  ae.su.toLocaleString());
      setEl('kpi_cvr_'+k2, cvr2+'%');
      setEl('kpi_cpa_'+k2, cpa2+'만');
    }});
    if (typeof adCpaChartRef !== 'undefined' && adCpaChartRef) {{
      adCpaChartRef.data.datasets[0].data = d.adEff.map(function(ae){{return ae.su?Math.round(ae.cost/ae.su):0;}});
      adCpaChartRef.update();
    }}
    if (typeof adEffChartRef !== 'undefined' && adEffChartRef) {{
      var maxC2 = Math.max.apply(null, d.adEff.map(function(ae){{return ae.cost;}}));
      d.adEff.forEach(function(ae, i) {{
        var cvr3 = ae.v?(ae.su/ae.v*100).toFixed(2):0;
        var cpa3 = ae.su?Math.round(ae.cost/ae.su/10000):0;
        adEffChartRef.data.datasets[i].data = [{{x:parseFloat(cvr3),y:cpa3,r:Math.max(8,Math.round(ae.cost/maxC2*28))}}];
      }});
      adEffChartRef.update();
    }}
  }}

  // rawData 테이블 업데이트
  if (d.rawRows && typeof renderTable === 'function') {{
    var newRawData = [];
    var structure2 = [
      {{cat:'direct',rows:['direct']}},
      {{cat:'unassigned',rows:['blog','outstanding','qletter','iboss','openads','stibee','etc.']}},
      {{cat:'paid search',rows:['brand_Search','google','naver','instagram']}},
      {{cat:'organic search',rows:['google','naver','etc.(bing)']}},
      {{cat:'referral',rows:['catenoid.net / charlla.io','cafe24 (ad)','makeshop','ai','etc.']}},
      {{cat:'display',rows:['GDN banner','CAFE24 banner','IBOSS banner','Outstanding banner','etc.(GFA, DMP..)']}},
      {{cat:'SNS',rows:['facebook / linkedin']}},
    ];
    structure2.forEach(function(s) {{
      var catRows = [];
      s.rows.forEach(function(rowName) {{
        var rd = (d.rawRows[s.cat]||{{}})[rowName] || {{v:0,ft:0,ac:0,po:0,su:0}};
        var cvr4 = rd.v&&rd.su ? (rd.su/rd.v*100).toFixed(1)+'%' : (rd.v?'0%':'-');
        var ct = rd.v>0?(rd.su/rd.v>0.07?'high':rd.su/rd.v>0.02?'mid':rd.su/rd.v>0?'low':'null'):'null';
        catRows.push({{name:rowName,v:rd.v,ft:rd.ft,ac:rd.ac,po:rd.po,su:rd.su,cvr:cvr4,cvrType:ct}});
      }});
      newRawData.push({{cat:s.cat,rows:catRows}});
    }});
    renderTable('rawTableBody', newRawData, {{high:'bg-emerald-100 text-emerald-700',mid:'bg-blue-100 text-blue-700',low:'bg-red-100 text-red-700',null:'bg-gray-100 text-gray-400'}});
    var allR = newRawData.flatMap(function(x){{return x.rows;}});
    setEl('accTotalV',  allR.reduce(function(s,r){{return s+r.v;}},0).toLocaleString());
    setEl('accTotalFt', allR.reduce(function(s,r){{return s+r.ft;}},0).toLocaleString());
    setEl('accTotalAc', allR.reduce(function(s,r){{return s+r.ac;}},0).toLocaleString());
    setEl('accTotalPo', allR.reduce(function(s,r){{return s+r.po;}},0).toLocaleString());
    setEl('accTotalSu', allR.reduce(function(s,r){{return s+r.su;}},0).toLocaleString());
  }}
}}
</script>'''

    return tabs_html, switch_script

tabs_html, switch_script = make_tab_html(months_data, cur_month_key)

# 탭 HTML을 헤더 바로 아래에 삽입
header_end = c.find('</header>') + len('</header>')
c = c[:header_end] + '\n' + tabs_html + c[header_end:]

# ── 차트 데이터 script 주입
mo = report_month
weekly_labels = [
    [f'{mo}월 1주차', f'({mo}/1~{mo}/8)'],
    [f'{mo}월 2주차', f'({mo}/9~{mo}/15)'],
    [f'{mo}월 3주차', f'({mo}/16~{mo}/22)'],
    [f'{mo}월 4주차', f'({mo}/23~{mo}/말)'],
]

adEff_cpa  = [d['cost']//d['su'] if d['su'] else 0 for d in adEff]
adEff_cvr_ = [round(d['su']/d['v']*100,2) if d['v'] else 0 for d in adEff]
adEff_cost = [d['cost'] for d in adEff]
maxCost    = max(adEff_cost) if adEff_cost else 1

inits  = 'window.addEventListener("load",function(){\n'
inits += 'if(window.weeklyAprilChartRef){var w=window.weeklyAprilChartRef;w.data.labels='+json.dumps(weekly_labels)+';w.data.datasets[0].data='+json.dumps(week_v)+';w.data.datasets[1].data='+json.dumps(week_mb)+';w.data.datasets[2].data='+json.dumps(week_ga4)+';w.update();}\n'
inits += 'if(window.monthlyTopChartRef){var mt=window.monthlyTopChartRef;mt.data.labels='+json.dumps(month_labels_list)+';mt.data.datasets[0].data='+json.dumps(monthly_v_list)+';mt.data.datasets[1].data='+json.dumps(monthly_mb_list)+';mt.data.datasets[2].data='+json.dumps(monthly_ga4_list)+';mt.update();}\n'
inits += 'if(window.cvrTrendChartRef){var ct=window.cvrTrendChartRef;ct.data.labels='+json.dumps(month_labels_list)+';ct.data.datasets[0].data='+json.dumps(monthly_mb_cvr_list)+';ct.data.datasets[1].data='+json.dumps(monthly_ga4_cvr_list)+';ct.update();}\n'
inits += 'if(window.paidConvChartRef){var pc=window.paidConvChartRef;pc.data.labels='+json.dumps(month_labels_list)+';pc.data.datasets[0].data='+json.dumps(monthly_paid_mb_list)+';pc.data.datasets[1].data='+json.dumps(monthly_paid_list)+';pc.data.datasets[2].data='+json.dumps(monthly_paid_r_list)+';pc.update();}\n'
inits += 'if(window.cvrChartRef){var cv=window.cvrChartRef;cv.data.labels='+json.dumps([d['name'] for d in cvrS])+';cv.data.datasets[0].data='+json.dumps([round(d['su']/d['v']*100,1) if d['v'] else 0 for d in cvrS])+';cv.data.datasets[0].backgroundColor='+json.dumps([d['color']+'bb' for d in cvrS])+';cv.data.datasets[0].borderColor='+json.dumps([d['color'] for d in cvrS])+';cvrSorted.length=0;'+json.dumps(cvrS)+'.forEach(function(d){cvrSorted.push(d);});cv.update();}\n'
for i in range(len(ch_data)):
    inits += 'if(window.channelTrendChartRef){window.channelTrendChartRef.data.labels='+json.dumps(month_labels_list)+';window.channelTrendChartRef.data.datasets['+str(i)+'].data='+json.dumps(ch_trend_pct[i])+';window.channelTrendChartRef.data.datasets['+str(i)+'].su='+json.dumps(ch_trend_abs[i])+';}\n'
inits += 'if(window.channelTrendChartRef){window.channelTrendChartRef.update();}\n'
inits += 'if(typeof adCpaChartRef!=="undefined"&&adCpaChartRef){adCpaChartRef.data.datasets[0].data='+json.dumps(adEff_cpa)+';adCpaChartRef.update();}\n'
inits += 'if(typeof adEffChartRef!=="undefined"&&adEffChartRef){'+\
    ''.join(['adEffChartRef.data.datasets['+str(i)+'].data=[{x:'+str(adEff_cvr_[i])+',y:'+str(round(adEff_cpa[i]/10000))+',r:'+str(max(8,round(adEff_cost[i]/maxCost*28)))+'}];' for i in range(len(adEff))])+\
    'adEffChartRef.update();}\n'
inits += '});\n'

c = c.rstrip()
if c.endswith('</body></html>'):
    c = c[:-len('</body></html>')] + '<script>\n' + inits + '</script>\n' + switch_script + '\n</body></html>'
else:
    c = c.replace('</html>', '<script>\n' + inits + '</script>\n' + switch_script + '\n</html>', 1)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(c)

print(f"완료: {cur_month_key} / 유입 {total_v:,}명 / GA4가입 {total_su}명 / MB가입 {mb_cur}명")
